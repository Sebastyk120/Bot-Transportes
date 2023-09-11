import locale
from datetime import datetime, timedelta
from time import sleep
import pyautogui
import webbrowser
from openpyxl import load_workbook
import pathlib
from pynput.keyboard import Controller

keyboard = Controller()
locale.setlocale(locale.LC_ALL, ("esp", "UTF-8"))
# ----------------------------------------------- Cargar TXT ----------------------------------------------------------
llave = dict()
nombre_archivo = 'Trasnportes_texto.txt'
link_grupo = 'https://web.whatsapp.com/accept?code=E2YLsn9pOxH52lBJZtPeC9'


def cargar_texto(llave, nombre_archivo):
    if pathlib.Path(nombre_archivo).exists():
        with open(nombre_archivo, 'r') as archivo:
            for linea in archivo:
                ides, hora_solicitud, tipo_de_servicio, fecha_servicio, hora_servicio, producto, cantidad_producto, kg, origen, destino, proveedor, celular_proveedor, conductor, celular_conductor, nombre_solicitud, estado = linea.strip().split(
                    ',')
                llave.setdefault(ides, (
                    hora_solicitud, tipo_de_servicio, fecha_servicio, hora_servicio, producto, cantidad_producto, kg,
                    origen, destino, proveedor, celular_proveedor, conductor, celular_conductor, nombre_solicitud,
                    estado))
    else:
        with open(nombre_archivo, 'w') as archivo:
            pass


# ----------------------------------------------- Cargar Excel y combinar con Texto-------------------------------------
# load excel file
cargar_texto(llave, nombre_archivo)


def cargar_excel(llave, nombre_archivo):
    workbook = load_workbook(filename="C:/Users/Sebas_2/HEAVEN/HeavensCorp - Documentos/Programación Transporte.xlsx")
    # open workbook
    sheet = workbook.active
    rango_condicion = sheet[f'A2:AB{sheet.max_row}']
    rango_ide = sheet[f'A{sheet.min_row}:A{sheet.max_row}']
    # ------------------------------------- Agregar al archivo de Texto los Ides que no existan ---------------------------
    for i in rango_condicion:
        clave = str(i[0].value)
        if llave.get(clave):
            print('Existe')
        else:
            ides = str(i[0].value)
            hora_solicitud = i[2].value + timedelta(hours=1)
            hora_solicitud = str(datetime.strftime(hora_solicitud, '%d-%B-%Y %H:%M:%S')).replace(',', '-')
            tipo_de_servicio = str(i[5].value.replace(',', '-'))
            fecha_servicio = str(datetime.strftime(i[7].value, '%d-%B-%Y')).replace(',', '-')
            hora_servicio = str(i[8].value).replace(',', '-')
            producto = str(i[9].value).replace(',', '-')
            cantidad_producto = str(i[10].value).replace(',', '-')
            kg = str(i[11].value).replace(',', '-')
            origen = str(i[12].value).replace(',', '-')
            destino = str(i[13].value).replace(',', '-')
            proveedor = str(i[14].value).replace(',', '-')
            celular_proveedor = str(i[15].value).replace(',', '-')
            conductor = str(i[16].value).replace(',', '-')
            celular_conductor = str(i[17].value).replace(',', '-')
            nombre_solicitud = str(i[27].value).replace(',', '-')
            estado = 'Enviado'
            webbrowser.open(link_grupo)
            sleep(10)
            keyboard.type('*Bot Heavens ------> Consecutivo:* ' + ides)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Tipo De Servicio:* ' + tipo_de_servicio)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Fecha Del Servicio:* ' + fecha_servicio)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Hora Del Servicio:* ' + hora_servicio)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Producto:* ' + producto)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Cantidad Producto:* ' + cantidad_producto)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Cantidad Kg:* ' + kg)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Origen:* ' + origen)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Destino:* ' + destino)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Proveedor:* ' + proveedor)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Celular Proveedor:* ' + celular_proveedor)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Conductor:* ' + conductor)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Celular Conductor:* ' + celular_conductor)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Solicitante:* ' + nombre_solicitud)
            with pyautogui.hold('shift'):
                pyautogui.press(['enter'])
            keyboard.type('*Hora Y Fecha De Creacion De Recogida:* ' + hora_solicitud)
            pyautogui.press('enter')
            sleep(8)
            pyautogui.hotkey('alt', 'f4')
            with open(nombre_archivo, 'a') as archivo:
                archivo.write(
                    f'{ides}, {hora_solicitud}, {tipo_de_servicio}, {fecha_servicio}, {hora_servicio}, {producto}, {cantidad_producto}, {kg}, {origen}, {destino}, {proveedor}, {celular_proveedor}, {conductor}, {celular_conductor}, {nombre_solicitud}, {estado}\n')
            print("La recogida se ha agregado con exito.")
            workbook.close()


cargar_excel(llave, nombre_archivo)
# ------------------------------------- Enviar mensajes en estado No enviado  -----------------------------------------
quit()
